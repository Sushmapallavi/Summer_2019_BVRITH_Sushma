import openpyxl
import click
import string
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
django.setup()
from django.conf import settings
from onlineapp.models import *


# settings.configure(DEBUG=True)
def extractDBname(string):
    return string.split('_')[2]

@click.command()
@click.argument('file')
def importFile(file):
    inpf = openpyxl.load_workbook(file)
    sheet = inpf["Colleges"]
    for row in sheet.iter_rows(min_row=2,values_only=True):
        clg = College(name=row[0],location=row[2],acronym=row[1],contact=row[3])
        clg.save()

    sheet = inpf["Current"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            clg = College.objects.get(acronym=row[1])
            stu = Student(name=row[0], email=row[2],db_folder=row[3].lower(), dropped_out=row[4],college=clg)
            stu.save()
        except clg.DoesNotExist:
            continue


    sheet = inpf["Deletions"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            clg = College.objects.get(acronym=row[1])
            stu = Student(name=row[0], email=row[2], db_folder=row[3].lower(), dropped_out=row[4], college=clg)
            stu.save()
        except clg.DoesNotExist:
            continue

    sheet = inpf["mock_test_results"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            stud = Student.objects.get(db_folder = extractDBname(row[0]))
            marks = MockTest1(problem1=row[1], problem2=row[2], problem3=row[3], problem4=row[4], total=row[5], student=stud)
            marks.save()
        except stud.DoesNotExist:
            continue


if __name__ == '__main__':
    importFile()