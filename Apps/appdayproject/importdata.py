import openpyxl
import click
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'appdayproject.settings')
django.setup()
from iplapp.models import *



@click.command()
@click.argument('file')
@click.argument('s')
def importFile(file,s):
    inpf = openpyxl.load_workbook(file)
    if s=="matches":
        sheet = inpf["matches"]
        for row in sheet.iter_rows(min_row=2,values_only=True):
            m = Matches(season=row[1],city=row[2],date=row[3],team_1=row[4],team_2=row[5],toss_winner=row[6],toss_decision=row[7],
                    result=row[8],dl_applied=row[9],winner=row[10],win_by_runs=row[11],win_by_wickets=row[12],player_of_match=row[13],
                    venue=row[14],umpire1=row[15],umpire2=row[16],umpire3=row[17])
            m.save()

    elif s=="deliveries":
        sheet = inpf["Sheet1"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            print(row[0])
            try:
                mat = Matches.objects.get(id=row[0])
                d = Deliveries(inning=row[1],batting_team=row[2],bowling_team=row[3],over=row[4],ball=row[5],batsman=row[6],non_striker=row[7],
                    bowler=row[8],is_super_over=row[9],wide_runs=row[10],bye_runs=row[11],legbye_runs=row[12],noball_runs=row[13],
                    penalty_runs=row[14],batsman_runs=row[15],extra_runs=row[16],total_runs=row[17],player_dismissed=row[18],dismissal_kind=row[19],fielder=row[20]
                    ,match=mat)
                d.save()
            except mat.DoesNotExist:
                continue


if __name__ == '__main__':
    importFile()